#!/usr/bin/python3
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.styles import Font
import requests
import csv
from bs4 import BeautifulSoup
from yahoofinancials import YahooFinancials
import yfinance as yf
import datetime
from datetime import date
from dateutil.relativedelta import relativedelta
import pandas as pd
from yahoo_fin import stock_info as si
from yahoo_earnings_calendar import YahooEarningsCalendar
import simplejson
import json
import numpy as np


import StockAnaPy




####### VARIABLES #######

# Index Tickers to track.
Index_Tickers = ["VOO", "QQQ", "VNQ", "VYM", "GC=F", "^TNX", "^FTSE", "^FCHI", "^GDAXI", "^N225"]

# Deltas for the price change.
Deltas = [1, 2, 3, 5, 10, 15, 30, 60, 90]

Data = 20
Delta_for_div = 200 # How many days back does it download the price data.
Yoy_delta = 4 # 4 quartes per year.
file = 'Analysis_SP500' ## FILE NAME ##
Sheet = 'Companies' ## SHEET NAME ##
Sheet2 = 'Underperforming Companies' ## Companies that fit the description of undervalued. ##
Index_sheet = 'Indices' ## Index Sheet Name
ran = 800 # How many columns total.

# Undervalued companies cutoffs
cutoff = -0.1 #Price Change
pe_cutoff = 15 #P/E
pb_cutoff = 3 # P/B
cr_cutoff = 1 # Current Ratio
de_cutoff = 2.5 # Debt/Equity Ratio


# Metrics to be updated.
combined_columns = ["Price"] + Deltas + ["Market Cap", "P/E Ratio", "P/E Ratio 3yravg.", "P/B Ratio", "Rev/MC", "Book/MC", "EPS at PE15", "EPS at PE20"]


for_industry_mc_metrics = Deltas + ["Market Cap"]

# Metrics to be looked at for undervalued companies.
Value_Metrics = ["P/E Ratio", "P/B Ratio", "Current Ratio", "Debt/Equity"]
Value = Deltas + Value_Metrics

# Metrics that need to be reused to calculate new ones.
recycled_metrics = ["Shares Outstanding", "EPS Yoy", "EPS 2018", "EPS 2019", "EPS 2020", "Book Value", "Revenue Yoy"]

# Load the workbook.
wb = load_workbook("/Users/jules/" + file + ".xlsx")
ws = wb[Sheet] 
ws_2 = wb[Sheet2]

# Main dataframe where everything gets copied to.
Main = pd.DataFrame(columns = combined_columns, index = range(ran))

# Dataframe of companies that are undervalued.
Value_Main = pd.DataFrame(columns = Value, index = range(ran))


# Where to start if bugs
company = 'PGR'
start_i = StockAnaPy.start_from_here_in_excel(ws, company) - 1

####### MAIN ########


# Create a list of column numbers for different metrics. These are helpful when creating new metrics efficiently.
new_column_numbers = StockAnaPy.find_column(ws, combined_columns)
recycled_column_numbers = StockAnaPy.find_column(ws, recycled_metrics)
for_industry_mc = StockAnaPy.find_column(ws, for_industry_mc_metrics)
delta_columns = StockAnaPy.find_column(ws, Deltas)
Value_columns = StockAnaPy.find_column(ws, Value_Metrics)

## Matrix for tops and bottoms of each metric.
#
#top_cutoff = 10
#
##Top_Bottom_Metrics_with_Values = StockAnaPy.add_value_columns(Top_Bottom_Metrics)
#
#Top_Main = pd.DataFrame(0, columns = Top_Bottom_Metrics_with_Values, index = range(top_cutoff))
#Bottom_Main = pd.DataFrame(0, columns = Top_Bottom_Metrics_with_Values, index = range(top_cutoff))

# Loop through every row to update the metrics
for row in range(2, ran+1) :

    # Start where the latest ticker was worked on.
    if row < start_i :
       continue

    # cell is the ticker.
    cell = ws.cell(row = row, column = 1).value
    
    # Skip row if no compnay in row.
    if cell == None :
        continue
    
    else :
        i = cell
        print ("Working on this stock",  i)
        if ws.cell(row = row, column = 3).value == None :
            ws.cell(row = row, column = 1).value = None
            continue

    # Populate the DataFrame with the Metrics to be used to create the new ratios aka. recycled metrics
    Recycled_Main = pd.DataFrame(columns = recycled_metrics, index = range(1))
    
    num = 0
    for metric in recycled_metrics :
        Recycled_Main.loc[0, metric] = ws.cell(row = row, column = recycled_column_numbers[num]).value
        num = num + 1
        
    # Function that returns the data delta as a string to be thrown into Yahoo Finance
    Time_frame = StockAnaPy.Timeframe(Delta_for_div)
    
    # Get historictical Prices for the Ticker
    data = yf.download(i, threads = False, start=str(Time_frame[1]), end=str(Time_frame[0]))
    
    # Gets the close prices.
    Low = data['Close']
    print (Low)
    Len = len(Low)
    
    # Get latest price for the Ticker
    live_price = StockAnaPy.Get_Live_Price_Yahoo(i)
    
    # If no new price, no need to update
    if live_price is False :
        continue
        
    else :
        Main.loc[row-1, "Price"] = live_price
    j_2 = 1
    
    for j in Deltas :
        try :
            print (Low[Len-j])
            Main.loc[row-1, j] = (((live_price / Low[Len-j]) - 1))
            
        except IndexError :
            break
            
        j_2 = j_2 + 1
        
    # Update the Market Cap.
    try :
        Market_cap = Recycled_Main.loc[0, "Shares Outstanding"] * live_price
        Main.loc[row-1, "Market Cap"] = Market_cap
        
    except TypeError :
        Main.loc[row-1, "Market Cap"] = "None"
        
    # Update the P/E Ratio.
    try :
        pe_ratio = live_price / Recycled_Main.loc[0, "EPS Yoy"]
        Main.loc[row-1, "P/E Ratio"] = pe_ratio
        
    except TypeError :
        Main.loc[row-1, "P/E Ratio"] = "None"
        
    except ZeroDivisionError:
        Main.loc[row-1, "Rev/MC"] = "None"
        
    # Update the P/B Ratio
    try :
        pb_ratio = Market_cap / Recycled_Main.loc[0, "Book Value"]
        Main.loc[row-1, "P/B Ratio"] = pb_ratio
        
    except TypeError :
        Main.loc[row-1, "P/B Ratio"] = "None"
        
    except ZeroDivisionError:
        Main.loc[row-1, "Rev/MC"] = "None"
        
    # Update the Revenue/Market Cap
    try :
        Main.loc[row-1, "Rev/MC"] = Recycled_Main.loc[0, "Revenue Yoy"] / Market_cap
        
    except TypeError :
        Main.loc[row-1, "Rev/MC"] = "None"
        
    except ZeroDivisionError:
        Main.loc[row-1, "Rev/MC"] = "None"
    
    # Update the Book/Market Cap
    try :
        Main.loc[row-1, "Book/MC"] = Recycled_Main.loc[0, "Book Value"] / Market_cap
        
    except TypeError :
        Main.loc[row-1, "Book/MC"] = "None"
        
    # Update the Avg. EPS.
    try :
        Avg_EPS = pd.Series.mean(Recycled_Main.loc[0, "EPS 2018":"EPS 2020"])
        Main.loc[row-1, "P/E Ratio 3yravg."] = live_price / Avg_EPS
        
    except TypeError :
        Main.loc[row-1, "P/E Ratio at 3yravg."] = "None"
    
    # Update the EPS at PE15
    try :
        Main.loc[row-1, "EPS at PE15"] = live_price / 15
        
    except TypeError :
        Main.loc[row-1, "EPS at PE15"] = "None"
        
    # Update the EPS at PE20
    try :
        Main.loc[row-1, "EPS at PE20"] = live_price / 20
        
    except TypeError :
        Main.loc[row-1, "EPS at PE20"] = "None"
        
    print (i, Main.loc[row-1, :])
    
    # Export all the new metrics to the excel row
    StockAnaPy.export_new_metrics_row_to_excel(wb, file, ws, new_column_numbers, Main, combined_columns, row-1)

print (Main)

wb.save("/Users/jules/" + file + ".xlsx")

#StockAnaPy.export_new_metrics_to_excel(ws, new_column_numbers, Main, combined_columns)


###### Here is the part where the averages for the Sub-Industries happen. #####
###############################################################################

Top_Bottom_Metrics_List = []

# Loop through all the rows, once it finds an empty row, it will start averaging the companies beneath that.
for row in range(2, ran+1) :

    # Skip row if no compnay in row.
    cell = ws.cell(row = row, column = 1).value
    if cell == None :
        continue
        
    else :
        i = cell
        print ("Working on this stock",  i)

    # Finds the start of the sub_industry list of companies
    if ws.cell(row = row - 1, column = 1).value == None :
    
        # Create some variables to add up all the values of the company metrics.
        live_market_cap = 0
        total_metric_with_weights = 0
        # A dataframe that stores the total market caps of all the companies in the sub-industry in the past at their respective deltas.
        past_market_caps = pd.DataFrame(0, index = range(1), columns = Deltas)
        row_1 = row

        # Loop until there is an empty row.
        while ws.cell(row = row_1, column = 1).value != None :
        
#            # Append all the metrics to the
#            metrics_in_row = StockAnaPy.append_list_of_metrics(ws, Top_Bottom_numbers, row_1)
#            Top_Bottom_Metrics_List.append(metrics_in_row)

            # Get the Live Market Cap Total
            live_market_cap += ws.cell(row = row_1, column = for_industry_mc[len(for_industry_mc)-1]).value

            past_col = 0
            past_col_2 = 0
            for d in Deltas :

                # You calculate the pre_mark_cap / (% + 1) to find the past_market cap
                delta_diff = ws.cell(row = row_1, column = for_industry_mc[past_col]).value
                if delta_diff is not None :
                    print (delta_diff)
                    change = ws.cell(row = row_1, column = for_industry_mc[len(for_industry_mc)-1]).value / (delta_diff + 1)

                else :
                    change = 0
                    ws.cell(row = row_1, column = for_industry_mc[past_col]).value = 0

                # change == None sometimes, End results resulted in None
                if change == None :
                    change = 0

                # Total past market caps get added up per Delta
                past_market_caps.loc[0, d] += change

                # To update the value metrics
                Value_Main.loc[row_1, d] = ws.cell(row = row_1, column = for_industry_mc[past_col]).value

                past_col = past_col + 1

            for m in Value_Metrics :
                print ("Excel Cell : ",  ws.cell(row = row_1, column = Value_columns[past_col_2]).value)
                Value_Main.loc[row_1, m] = ws.cell(row = row_1, column = Value_columns[past_col_2]).value
                past_col_2 = past_col_2 + 1

            row_1 = row_1 + 1

        # Find the average of the value metrics

        # PE Ratio
        StockAnaPy.calculate_metric_total_for_sub_industry(ws, row, Value_columns[0], for_industry_mc, live_market_cap, 1)

        # PB Ratio
        StockAnaPy.calculate_metric_total_for_sub_industry(ws, row, Value_columns[1], for_industry_mc, live_market_cap, 0)

        # Current Ratio
        StockAnaPy.calculate_metric_total_for_sub_industry(ws, row, Value_columns[2], for_industry_mc, live_market_cap, 0)

        # Debt Equity
        StockAnaPy.calculate_metric_total_for_sub_industry(ws, row, Value_columns[3], for_industry_mc, live_market_cap, 0)

        # Find the average of delta differences with their market caps as a weight. The bigger the market cap, the more it will affect the overall delta difference.
        dx = 0
        for dd in Deltas :
            
            # Updates the past market cap adding the company cap everytime.
            past_market_caps.loc[0, dd] = live_market_cap/past_market_caps.loc[0, dd] - 1
            ws.cell(row = row - 1, column = delta_columns[dx]).value = past_market_caps.loc[0, dd]
            ws.cell(row = row - 1, column = delta_columns[dx]).font = Font(bold=True)
            dx = dx + 1

        ws.cell(row = row - 1, column = for_industry_mc[len(for_industry_mc)-1]).value = live_market_cap
        ws.cell(row = row - 1, column = for_industry_mc[len(for_industry_mc)-1]).font = Font(bold=True)
        ws.cell(row = row - 1, column = 2).font = Font(bold=True)

print (Top_Bottom_Metrics_List)

#Top_Bottom_Metrics_Array = np.array(Top_Bottom_Metrics_List)
#
#StockAnaPy.find_top_performers(Top_Bottom_Metrics_Array, Top_Bottom_Metrics, Top_Main, top_cutoff, Zeroish_is_Top)
#
#StockAnaPy.export_to_excel(file, Top_Bottom_Sheet, Top_Main, 3, 3)

# Top Gainers and Top Losers.
#




# Look for the undervalued companies.
new_row_num = 2
for i in range(2, ran) :
    print (Value_Main.loc[i, :])
    try :
        # Loops through every company, if its beneath all the cutoffs, then the row gets copied to the different sheet.
        if 0 < float(Value_Main.loc[i, "P/E Ratio"]) < pe_cutoff and 0 < float(Value_Main.loc[i, "P/B Ratio"]) < pb_cutoff and float(Value_Main.loc[i, "Current Ratio"]) > cr_cutoff :
            print ("This One")
            StockAnaPy.copy_row(ws, ws_2, i, new_row_num)
            new_row_num = new_row_num + 1

    except KeyError :
        print ("Here is the Error", ws.cell(row = i, column = 1).value)
        
    except TypeError :
        print ("Here is the Error", ws.cell(row = i, column = 1).value)
        
wb.save("/Users/jules/" + file + ".xlsx")






