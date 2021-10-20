#!/usr/bin/python3
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.styles import Font
import requests
import csv
from bs4 import BeautifulSoup
from yahoofinancials import YahooFinancials
import yfinance as yf
import datetime
from datetime import date
from dateutil.relativedelta import relativedelta
import pandas as pd
from yahoo_fin import stock_info as si
from yahoo_earnings_calendar import YahooEarningsCalendar
import textwrap
import math
import numpy as np



######## - - - - FUNCTIONS - - - - #########
############################################

# Returns HTML code using BeautifulSoup
def get_html_code(pg) :

    page = requests.get(pg)
    soup = BeautifulSoup(page.content, 'lxml')
    return soup # Returns the entire HTML code


# Returns the html code that is linked to the accounting principle
def find_string(soup, string) :

    table = soup.find_all('td')
    test = 0
    for t in table :
        if string in t.text :
            words = t.text.split()
            words = words[0:int((len(words)/2))]
            Words = string.split()
            if Words == words :
                test = 1
                break
                
    if test == 0 :
        return ("None")
        
    else :
        return t


# Returns the 5 datapoints for the accounting principle
def find_data(t) :
    
    Find_data = t.find_next_siblings('td')
    Data = []
    for data in Find_data :
        Data.append(data.text)
        
    return Data[0:len(Data)-1]


# Function to find the number from a string in Marketwatch
def find_num(string, l, v) :

    if "(" in string :
        str = string.replace("(", "-").replace(")", "")
        return float(str.replace(l, "")) * v
    else :
        return float(string.replace(l, "")) * v

# Function that checks if a comma is in a string. If so, it replaces it with ""
def check_comma(string) :
    
    if "," in string :
        str = string.replace(",", "")
        return str
    else :
        return string


# Converts Markwatch text into number if possible
def transform_to_num(strings) :
    
    new_strings = []
    for string in strings :
        str = check_comma(string)
        if "-" in str :
            new_strings.append(0)
            continue
            
        if "()" in str :
            new_strings.append(0)
            continue
            
        if "T" in str :
            number = find_num(str, "T", 1000000000000)
            new_strings.append(number)
            continue
            
        if "B" in str :
            number = find_num(str, "B", 1000000000)
            new_strings.append(number)
            continue
            
        if "M" in str :
            number = find_num(str, "M", 1000000)
            new_strings.append(number)
            continue
            
        if "K" in str :
            number = find_num(str, "K", 1000)
            new_strings.append(number)
            continue
            
        else :
            number = find_num(str, "", 1)
            new_strings.append(number)

    return new_strings


# returns Dates to be plugged in looking for the Delta
def Timeframe(delta) :

    past_datetime = datetime.datetime.today() - datetime.timedelta(days=delta)
    dates = [str(datetime.datetime.today()), str(past_datetime)]
    new_dates = []
    for d in dates :
        split_string = d.split(" ", 1)
        new_date = split_string[0]
        new_dates.append(new_date)
    print (new_dates)
    return new_dates


# Function to find the Yoy dividend yield. You can switch the delta. Price taken is the one at the start of the time frame.
def Yoy_div_yield(Div, delta, Low) :
    
    divi = 0
    past_date = (datetime.datetime.today() - datetime.timedelta(days=delta))
    data = Low[0]
    Dividend = Div["dividend"]
    print ("Here is the price from a year ago", data)
    for i in range(1, delta+1) :
        if Div.index[len(Div)-i] > past_date :
            print (Dividend[len(Div)-i])
            divi = divi + float(Dividend[len(Div)-i])
            print (divi)
        else :
            break
    return (divi / data) * 100


# Function to find the next earnings date
def next_earnings_date(i) :
    
    yec = YahooEarningsCalendar()
    earnings_list = yec.get_earnings_of(i)
    earnings_df = pd.DataFrame(earnings_list)
    try :
        Earnings_dates = earnings_df["startdatetime"]
    except KeyError :
        return "-"
    date = "-"
    for i in Earnings_dates :
        
        int_date = i.split("T")[0]
        cut_date = int_date.split("-")
        future_date = datetime.datetime(int(cut_date[0]), int(cut_date[1]), int(cut_date[2]))
        
        if future_date > datetime.datetime.today() :
            date = str(future_date).split(" ")[0]
        
        else :
            break

    return str(date)

# Function that returns a list with Yoy as [0] and yearlies depending on the delta_year
def adding_data_past_data(Sub_main, Sub_main_2, principle, og_col, delta_year, delta_quarter, year) :
    
    row = []
    sum = 0
    list = Sub_main_2.loc[principle, (og_col-3):og_col]
    for i in list :
        try :
            sum = sum + i
            
        except TypeError :
            continue

    row.append(sum)
    list_3 = Sub_main_2.loc[principle, og_col-delta_quarter+1:og_col]
    for i in list_3 :
        row.append(i)
    
    list_2 = Sub_main.loc[principle, (year-delta_year+1):year]
    for i in list_2 :
        row.append(i)
    
    return row


# Function to add to Row
def adding_to_Row(Row, list) :

    for i in list :

        Row.append(i)


# Function to export to excel. R and C are the first cell where the data will be placed.
def export_to_excel(file, page, M, r, c) :
    
    file_path = "/Users/jules/" + file + ".xlsx"
    wb=load_workbook(file_path)
    ws = wb[page]
    k = r
    
    for i in range(0, M.shape[0]) :
        
        ws.cell(row = k, column = r-1).value = M.index[i]
        
        for j in range(c, M.shape[1]+c) :
            
            ws.cell(row = k, column = j).value = M.iloc[i, j-c]
        
        k = k + 1

    for l in range(c, M.shape[1]+c) :
        
        ws.cell(row = r-1, column = l).value = M.columns[l-c]

    wb.save(file_path)
    print ("DONE!")




# Make sure that the Marketwatch year is the right one in the Annual report section.
def check_years_table(soup, x, year) :
    
    if x == 1 :
        
        th = soup.find_all(class_="overflow__heading")
        
        if th[len(th)-2].text == year :
            
            return 4
        
        else :
            
            return 3

    else :
    
        return 4


# Function to export list to excel. R and C are the first cell where the data will be placed.
def export_list_to_excel(file, page, M, r, c, row) :
    
    file_path = "/Users/jules/" + file + ".xlsx"
    wb=load_workbook(file_path)
    ws = wb[page]
    if row == 0 :
        
        for i in range(0, len(M)) :
            
            ws.cell(row = i+r, column = c).value = M[i]

    else :
    
        for i in range(0, len(M)) :
        
            ws.cell(row = r, column = i+c).value = M[i]
    
    wb.save(file_path)
    print (page, "is done")


# Function to find and return the column number for a list of metrics
def find_column(ws, column_names) :
    
    column_numbers = []
    i = 1
    while ws.cell(row = 1, column = i).value != None :
    
        if ws.cell(row = 1, column = i).value in column_names :
        
            column_numbers.append(i)

        i = i + 1

    return column_numbers


# Exports the list into the right columns
def export_new_metrics_to_excel(ws, new_column_numbers, Main, combined_columns) :

    for ticker in range(2, Main.shape[0]) :
        
        num = 0
        for col in new_column_numbers :
            print(col)
            ws.cell(row = ticker+1, column = col).value = Main.loc[ticker, combined_columns[num]]
            num = num + 1


def save_at_interval(wb, file, interval, row) :

    if row % interval == 0 :

        wb.save("/Users/jules/" + file + ".xlsx")
        print ("\n\n\nSAVED\n\n\n")

# Function that uploads each stock every time. Per row.
def export_new_metrics_row_to_excel(wb, file, ws, new_column_numbers, Main, combined_columns, ticker) :

    num = 0
    for col in new_column_numbers :
        print(col)
        ws.cell(row = ticker+1, column = col).value = Main.loc[ticker, combined_columns[num]]
        num = num + 1
    save_at_interval(wb, file, 25, ticker)

# Function to copy one row to another
def copy_row(ws, ws_2, old_row_num, new_row_num) :

    col = 1
    
    while ws.cell(row = old_row_num, column = col).value != None :

        ws_2.cell(row = new_row_num, column = col).value = ws.cell(row = old_row_num, column = col).value

        col = col + 1

# Parse Yahoo Finance live price.
def Get_Live_Price_Yahoo(ticker) :

    pg = "https://finance.yahoo.com/quote/" + ticker + "?p=" + ticker + "MMM&.tsrc=fin-srch"
    soup = get_html_code(pg)
    
    # Does not always download, so in that case, take the latest available price.
    try :
        price = soup.find(class_="Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)").text
        price = price.replace(',', '')
        return float(price)
        
    except AttributeError :
    
        return False
        


# Function to get quotes for all indices in ticker list.
def get_indices_quotes(Main, Tickers, Deltas) :
    
    Delta_for_div = 365
    Prices = pd.DataFrame(columns = Deltas, index = Tickers)
    
    for i in Tickers :
        
        
        Row = []
        changes = []
        
        # Function that returns the data delta as a string to be thrown into Yahoo Finance
        Time_frame = Timeframe(Delta_for_div)
        # Get historictical Prices for the Ticker
        data = yf.download(i, start=str(Time_frame[1]), end=str(Time_frame[0]))
        # Get latest price for the Ticker
        live_price = Get_Live_Price_Yahoo(i)
        Row.append(live_price)
        Low = data['Close']
        print (Low)
        Len = len(Low)
        
        for j in Deltas :
            
            print (Low[Len-j])
            changes.append((((live_price / Low[Len-j]) - 1) * 100))
            Row.append((((live_price / Low[Len-j]) - 1)))
        
        Prices.loc[i:i, :] = changes
        print (i, Row)
    
        Main.loc[i:i, :] = Row
        print(Main.loc[i:i, :])


def start_from_here_in_excel(ws, company) :

    i = 1
    while ws.cell(row = i, column = 1).value != company :
    
        i = i + 1

    return i

# Function that tells the program where to start from when mass data producing. This allows to not have to restart from 0 everytime.
def start_from_here(tickers, company_to_start_from) :

    i = 0
    while tickers.loc[i, 0] != company_to_start_from :
    
        i = i + 1

    return i

# Function to calculate the averages of metrics by Sub-Industry.
def calculate_metric_total_for_sub_industry(ws, start_row, col, for_industry_mc, live_market_cap, negative_values) :

    total = 0
    row_1 = start_row
    market_cap = 0
    
    if negative_values == 0 :
    
        market_cap = live_market_cap

    while ws.cell(row = row_1, column = 1).value != None :
        
        if ws.cell(row = row_1, column = col).value is None :
        
            row_1 += 1
            continue
        
        if ws.cell(row = row_1, column = col).value < 0 :
            
            row_1 += 1
            continue

        if negative_values == 1 :
            
            

            market_cap += ws.cell(row = row_1, column = for_industry_mc[len(for_industry_mc)-1]).value

        total += (ws.cell(row = row_1, column = col).value * ws.cell(row = row_1, column = for_industry_mc[len(for_industry_mc)-1]).value)

        row_1 += 1
    
    ws.cell(row = start_row - 1, column = col).font = Font(bold=True)
    
    print (market_cap)

    try :
        ws.cell(row = start_row - 1, column = col).value = total / market_cap
    except ZeroDivisionError :
        ws.cell(row = start_row - 1, column = col).value = 0

# Function that checks if the latest year has the number of outstanding shares.
def CheckLatestShares(value) :

    if value != 0 :
        return value

    else :
        return False

# Function to grab the previous amount of shares. From the past year.
def GetLatestShares(Sub_Main, metric, latest_year) :

    for x in range(1, latest_year) :
        if Sub_Main.loc[metric, latest_year - x] != 0 :
            return Sub_Main.loc[metric, latest_year - x]

    print ("\n\nNO SHARES\n\n")
    return "N/A"
    

# Add header to the top of every sheet in the excel sheets
def add_header(file, sheet, header) :
    
    file_path = "/Users/jules/" + file + ".xlsx"
    wb = load_workbook(file_path)
    ws = wb[sheet]
    
    for i in range(len(header)) :
        ws.cell(row = 1, column = i+2).value = header[i]
    
    wb.save(file_path)

# Append Metrics to be looked at for performance to a list.
def append_list_of_metrics(ws, Top_Bottom_numbers, Row) :

    list = []
    
    for metric in Top_Bottom_numbers :
        list.append(ws.cell(row = Row, column = metric).value)
        
    return list
    
# Function that adds a value column to each metric.
def add_value_columns(Metrics) :

    Metrics_with_Values = []
    
    for metric in Metrics :
    
        if metric is "Companies" :
        
            continue
    
        Metrics_with_Values.append(str(metric))
        print (metric)
        Metrics_with_Values.append(str(metric) + " Value")
        
    return Metrics_with_Values
    
# Function returns N largest elements
def N_max_elements(list, N, name_list) :

    final_list = pd.DataFrame(columns = range(2), index = range(N))
    final_list.loc[:, 1] = 0

    for i in range(0, len(list)) :
    
        if list[i] is None :
        
            continue
            
        if list[i] == 'None' :
        
            continue
            
        if float(list[i]) > float(final_list.loc[N-1, 1]) :

            Top = True

            for x in range(2, N+1) :
            
                if float(list[i]) < float(final_list.loc[N-x, 1]) :

                    final_list.loc[N-x+1, 0] = name_list[i]
                    final_list.loc[N-x+1, 1] = list[i]
                    Top = False
                    break
                    
                else :
                
                    final_list.loc[N-x+1, :] = final_list.loc[N-x, :]

            if Top is True :

                final_list.loc[0, 1] = list[i]
                final_list.loc[0, 0] = name_list[i]

    return final_list
    
# Function returns N smallest elements
def N_min_elements(list, N, name_list) :

    final_list = pd.DataFrame(columns = range(2), index = range(N))
    final_list.loc[:, 1] = 100000000000000000

    for i in range(0, len(list)) :
    
        if list[i] is None :
        
            continue
            
        if list[i] == 'None' :
        
            continue
            
        if list[i] <= 0 :
        
            continue
            
        if float(list[i]) < float(final_list.loc[N-1, 1]) :

            Top = True

            for x in range(2, N+1) :
            
                if float(list[i]) > float(final_list.loc[N-x, 1]) :

                    final_list.loc[N-x+1, 0] = name_list[i]
                    final_list.loc[N-x+1, 1] = list[i]
                    Top = False
                    break
                    
                else :
                
                    final_list.loc[N-x+1, :] = final_list.loc[N-x, :]

            if Top is True :

                final_list.loc[0, 1] = list[i]
                final_list.loc[0, 0] = name_list[i]

    return final_list

## Function to loop through different columns in list to find top fives. Put them all in a matrix.
def find_top_performers(array, Metrics, df, cutoff, Zeroish_is_Top) :

    name_list = array[:, 0]
    print (name_list)
    print (Metrics)
    print (array.shape)

    for col in range(1, len(Metrics)) :
    
        list = array[:, col]
    
        if Metrics[col] in Zeroish_is_Top :
            
            print (Metrics[col], "Is Zeroish")
            final_list = N_min_elements(list, cutoff, name_list)
            
        else :
        
            final_list = N_max_elements(list, cutoff, name_list)
            
        print (Metrics[col])
        print (final_list)
        
        for i in range(cutoff) :
        
            df.iloc[i, (col-1)*2] = final_list.iloc[i, 0]
            df.iloc[i, ((col-1)*2)+1] = final_list.iloc[i, 1]
        
        col += 1
        
    print (df)
        


# Clears all the data in the excel file.
def clear_excel_cells(file, page) :
    
    file_path = "/Users/jules/" + file + ".xlsx"
    wb = load_workbook(file_path)
    ws = wb[page]
    print (page)
    i = 2
    
    while ws.cell(row = i, column = 1).value != None :
            
        j = 1
        while ws.cell(row = 1, column = j).value != None :
                
            ws.cell(row = i, column = j).value = None
            j = j + 1
            
        i = i + 1

    wb.save(file_path)
    

# Function that returns a DataFrame of the current S&P 500 companies with their Sub-Industries from Wikipedia.
def get_SP500_companies(exceptions, i_2) :
    
    payload = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')
    table = payload[0]
    companies = table.shape[0]
    Main = pd.DataFrame(columns = range(3), index = range(companies-i_2))
    
    for i in range(i_2, companies) :
        
        # Check if the company is not on the exception list (cannot download MarketWatch data).
        if table.loc[i, "Symbol"] in exceptions :
            continue
    
        else :
            Main.loc[i, 0] = table.loc[i, "Symbol"]
            Sub_industry = table.loc[i, "GICS Sub-Industry"]
            
            # Excel only takes 31 characters for the name of the worksheets
            if len(Sub_industry) > 31 :
                Main.loc[i, 1] = Sub_industry[:31].lower()
            
            else :
                
                Main.loc[i, 1] = Sub_industry.lower()
            
            Main.loc[i, 2] = table.loc[i, "Security"]

    return Main.loc[:, :]
    
    
# Function that classifies the tickers in the right sheet of excel (in their Sub-Industry)
def classify_by_sheet(Main, column, file, header) :

    industries = []
    file_path = "/Users/jules/" + file + ".xlsx"
    wb = load_workbook(file_path)
    
    for i in range(Main.shape[0]) :
        if Main.loc[i, column] in industries :
            continue
    
        else :
            industry = Main.loc[i, column]
            industries.append(industry)
            sub_main = []
            for j in range(i, Main.shape[0]) :
                if industry == Main.loc[j, column] :
                    sub_main.append(Main.loc[j, 0])
                        
            if industry.lower() in (name.lower() for name in wb.sheetnames) :
                export_list_to_excel(file, industry, sub_main, 2, 1, 0)
                add_header(wb, industry, header, file_path)
            
            else :
                add_sheet(file, industry)
                add_header(file_path, industry, header, file_path)
                export_list_to_excel(file, industry, sub_main, 2, 1, 0)
        
        
# Function that looks where the acounting principle is and finds the string of the principle.
def which_statement_is_it_in(j, Accounting_Principles_Income, Accounting_Principles_Balance, Accounting_Principles_Flow, soup_income, soup_balancesheet, soup_cashflow) :
    
    if j in Accounting_Principles_Income :
        t = find_string(soup_income, j)
        return t
        
    if j in Accounting_Principles_Balance :
        t = find_string(soup_balancesheet, j)
        return t
        
    if j in Accounting_Principles_Flow :
        t = find_string(soup_cashflow, j)
        return t
        
# Function to check if the latest column in marketwatch actually has data. If not, it will take it from the previous column. Returns the column number to parse.
def complete_column(Sub_main_2, Accounting_Principles) :
    
    f = 0
    for metric in Accounting_Principles :
        if Sub_main_2.loc[metric, 4] != 0 :
            f = f + 1

    if f == 0 :
        return 3
    
    else :
        return 4

# Find the date of the accounting principle data in the html code.
def find_statement_date(soup, col) :
    
    th = soup.find_all(class_="overflow__heading")
    print (th[col+1].text)
    return th[col+1].text
    
# Function to create a list out of the numbers in the row of that ticker in main.
def from_Main_to_Row(Main, i, combined_columns) :
    
    Row = []
    for ro in combined_columns :
        Row.append(Main.loc[i, ro])
    
    return Row

# Find the row number of the ticker in the Worksheet to then use that number to place the row
def find_row_in_xcel(wb, industry, ticker) :
    
    ws = wb[industry]
    i = 2
    
    cell = ws.cell(row = i, column = 1).value
    while cell != None :
        
        if cell == ticker :
            
            return i
        
        i = i + 1
        cell = ws.cell(row = i, column = 1).value
    
    
    return 0


# Function that classifies the tickers and their metrics in the right sheet of excel
def classify_into_sheet(Row, industry, file, ticker) :
    
    file_path = "/Users/jules/" + file + ".xlsx"
    wb = load_workbook(file_path)
    xcel_row = find_row_in_xcel(wb, industry, ticker)
    export_list_to_excel(file, industry, Row, xcel_row, 2, 1)


# Function that puts all the data into the Sheet1 of excel.
def put_into_sheet1(file, sheet, new_file) :
    
    file_path = "/Users/jules/" + file + ".xlsx"
    new_file_path = "/Users/jules/" + new_file + ".xlsx"
    wb = load_workbook(file_path)
    new_wb = load_workbook(new_file_path)
    Main_sheet = new_wb[sheet]
    Main_row_count = 2
    
    
    for ws in wb.worksheets :
        
        if ws.title == "Sheet1" :
            print("SKIPT")
            continue
        
        ws.cell(row = 1, column = 1).value = "="
        print (ws.title)
        i = 2
        print (Main_row_count)
        Main_sheet.cell(row = Main_row_count + 1, column = 2).value = ws.title
        Main_sheet.cell(row = Main_row_count + 1, column = 2).font = Font(bold=True)
        
        while ws.cell(row = i, column = 1).value != None :
            
            j = 1
            
            while ws.cell(row = 1, column = j).value != None :
                
                Main_sheet.cell(row = Main_row_count + i, column = j).value = ws.cell(row = i, column = j).value
                j = j + 1
            
            i = i + 1
        
        Main_row_count = Main_row_count + i
    
    wb.save(file_path)
    new_wb.save(new_file_path)
